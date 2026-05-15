import streamlit as st
import pandas as pd
import os
import re
import io
import xml.etree.ElementTree as ET
from datetime import datetime
from auth import require_login

_PRIMARY = r"C:\Users\herru\OneDrive\Escritorio\Drive-Compartido\CONTROL DE FACTURAS EMITIDAS NETSAT 2026-2025.xlsx"
_BACKUP  = "CONTROL DE FACTURAS EMITIDAS NETSAT 2026-2025.xlsx"
_XML_DIR = r"Z:\NETSAT\NETSAT 2026\FACTURAS GUIAS NETSAT SRL 2026\FACTURAS VENTAS 2026\XML 2026"

_NS = {
    "cbc": "urn:oasis:names:specification:ubl:schema:xsd:CommonBasicComponents-2",
    "cac": "urn:oasis:names:specification:ubl:schema:xsd:CommonAggregateComponents-2",
}

def _txt(el, path):
    node = el.find(path, _NS)
    return node.text.strip() if node is not None and node.text else ""

def _path():
    return _PRIMARY if os.path.exists(_PRIMARY) else _BACKUP

def _usd(n):
    if n >= 1_000_000: return f"{n/1_000_000:.1f}M"
    if n >= 1_000:     return f"{n/1_000:.1f}k"
    return f"{n:.0f}"

def _find_header(df, keywords):
    """Devuelve el df con encabezados a partir de la fila que contiene alguna keyword."""
    for i, row in df.iterrows():
        if any(any(k in str(v).upper() for k in keywords) for v in row if pd.notna(v)):
            df.columns = df.iloc[i]
            df = df.iloc[i + 1:].reset_index(drop=True)
            df.columns = [str(c).strip() for c in df.columns]
            return df.dropna(how="all")
    return df


# ─── CARGA ─────────────────────────────────────────────────────────────────────

@st.cache_data(ttl=0)
def cargar_proyectos():
    df = pd.read_excel(_path(), sheet_name="PROY-2026", header=None)
    df = _find_header(df, ["#PROYECTO", "O/C"])
    df = df.rename(columns={
        "#Proyecto": "Proyecto",
        "O/C": "OC_raw",
        "Fac. Ventas": "Factura_Venta",
        "Guía Ventas": "Guia_Venta",
        "Fac. Compras": "Factura_Compra",
        "Inscripción": "Fecha_Inicio",
        "Estimada Fin": "Fecha_Fin",
    })
    df["OC_base"] = df["OC_raw"].astype(str).str.extract(r"(C\d+)")[0]
    for col in ["Fecha_Inicio", "Fecha_Fin"]:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors="coerce")
    return df


@st.cache_data(ttl=0)
def cargar_guias():
    df = pd.read_excel(_path(), sheet_name="GUIAS", header=None)
    df = _find_header(df, ["GUIA", "ESTATUS"])
    df.columns = ["Guia", "Fecha", "OC", "Estatus", "Factura", "Detalle"]
    df["Fecha"]    = pd.to_datetime(df["Fecha"], errors="coerce")
    df["OC"]       = df["OC"].astype(str)
    df["Factura"]  = df["Factura"].astype(str)
    df["Guia_num"] = df["Guia"].astype(str).str.extract(r"T001-(\d+)")[0].astype(float)
    return df.sort_values("Guia_num").reset_index(drop=True)


@st.cache_data(ttl=0)
def cargar_coupa():
    df = pd.read_excel(_path(), sheet_name="COUPA2026", header=None)
    df = _find_header(df, ["FACTURA", "O/C", "COUPA"])
    df.columns = ["Factura", "Dig", "Fecha", "Monto", "Dias_Credito", "Fecha_Pago_Pactada", "OC", "COUPA", "Fecha_Pago_Real", "Pago"]
    df["Fecha"]              = pd.to_datetime(df["Fecha"],              errors="coerce")
    df["Fecha_Pago_Pactada"] = pd.to_datetime(df["Fecha_Pago_Pactada"], errors="coerce")
    df["Fecha_Pago_Real"]    = pd.to_datetime(df["Fecha_Pago_Real"],    errors="coerce")
    df["Monto"]              = pd.to_numeric(df["Monto"], errors="coerce").fillna(0)
    df["OC"]                 = df["OC"].astype(str)
    df["Pago"]               = df["Pago"].astype(str).str.strip()
    return df


@st.cache_data(ttl=0)
def cargar_facturas():
    df = pd.read_excel(_path(), sheet_name="FACTURAS", header=None)
    df = _find_header(df, ["FACTURA", "CLIENTE"])
    df.columns = [
        "Factura", "Fecha", "Cliente",
        "Val_Sol", "IGV_Sol", "Total_Sol",
        "Val_Dol", "IGV_Dol", "Total_Dol",
        "Pagado", "Retencion", "Detraccion",
        "Guia", "OC", "Observaciones",
    ]
    df["Fecha"]     = pd.to_datetime(df["Fecha"], errors="coerce")
    df["Total_Dol"] = pd.to_numeric(df["Total_Dol"], errors="coerce").fillna(0)
    df["Total_Sol"] = pd.to_numeric(df["Total_Sol"], errors="coerce").fillna(0)
    df["OC"]        = df["OC"].astype(str)
    df["Guia"]      = df["Guia"].astype(str)
    return df


@st.cache_data(ttl=0)
def parsear_xmls():
    """Lee todos los XMLs de la carpeta y devuelve un DataFrame con los datos extraídos."""
    if not os.path.isdir(_XML_DIR):
        return pd.DataFrame()

    registros = []
    for raiz_dir, _, archivos in os.walk(_XML_DIR):
        for nombre in archivos:
            if not nombre.upper().endswith(".XML"):
                continue
            ruta = os.path.join(raiz_dir, nombre)
        try:
            tree = ET.parse(ruta)
            raiz = tree.getroot()

            factura   = _txt(raiz, "cbc:ID")
            fecha_str = _txt(raiz, "cbc:IssueDate")
            moneda    = _txt(raiz, "cbc:DocumentCurrencyCode")

            cliente = (
                _txt(raiz, "cac:AccountingCustomerParty/cac:Party/cac:PartyLegalEntity/cbc:RegistrationName")
                or _txt(raiz, "cac:AccountingCustomerParty/cac:Party/cac:PartyName/cbc:Name")
            )
            cliente = cliente.upper()
            if "ANTAPACCAY" in cliente:
                cliente = "ANTAPACCAY"
            elif "SUNAT" in cliente:
                cliente = "SUNAT"

            oc = _txt(raiz, "cac:OrderReference/cbc:ID")

            val_sin_igv = float(_txt(raiz, "cac:LegalMonetaryTotal/cbc:LineExtensionAmount") or 0)
            igv         = float(_txt(raiz, "cac:TaxTotal/cbc:TaxAmount") or 0)
            total       = float(_txt(raiz, "cac:LegalMonetaryTotal/cbc:PayableAmount") or 0)

            if moneda == "USD":
                val_sol, igv_sol, total_sol = 0.0, 0.0, 0.0
                val_dol, igv_dol, total_dol = val_sin_igv, igv, total
            else:
                val_sol, igv_sol, total_sol = val_sin_igv, igv, total
                val_dol, igv_dol, total_dol = 0.0, 0.0, 0.0

            fecha = pd.to_datetime(fecha_str, errors="coerce")

            registros.append({
                "Factura":   factura,
                "Fecha":     fecha,
                "Cliente":   cliente,
                "Val_Sol":   val_sol,
                "IGV_Sol":   igv_sol,
                "Total_Sol": total_sol,
                "Val_Dol":   val_dol,
                "IGV_Dol":   igv_dol,
                "Total_Dol": total_dol,
                "OC":        oc,
                "Moneda":    moneda,
                "Archivo":   nombre,
            })
        except Exception:
            pass

    if not registros:
        return pd.DataFrame()

    df = pd.DataFrame(registros)
    df["Fac_num"] = df["Factura"].str.extract(r"F\d+-(\d+)")[0].astype(float)
    return df.sort_values("Fac_num").reset_index(drop=True)


def _sugerencias_excel(df_sug):
    """Genera un .xlsx con el mismo formato de la hoja FACTURAS de Nélida."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SUGERENCIAS"

    encabezados = [
        "FACTURA N°", "FECHA EMISION", "CLIENTE",
        "VALOR SIN IGV SOL", "IGV SOLES", "TOTAL SOLES",
        "VALOR SIN IGV DÓLAR", "IGV DÓLAR", "TOTAL DÓLAR",
        "PAGADO", "RETENCION ENTREGADA", "DETRACCION ENTREGADA",
        "GUIA REMISION", "ORDEN DE COMPRA", "OBSERVACIONES",
    ]
    fill_header = PatternFill("solid", fgColor="1F4E79")
    font_header = Font(bold=True, color="FFFFFF", size=10)
    borde = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )

    for col_idx, titulo in enumerate(encabezados, start=1):
        cell = ws.cell(row=1, column=col_idx, value=titulo)
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = borde

    fill_fila = PatternFill("solid", fgColor="E2EFDA")
    for row_idx, (_, fila) in enumerate(df_sug.iterrows(), start=2):
        valores = [
            fila["Factura"],
            fila["Fecha"].date() if pd.notna(fila["Fecha"]) else None,
            fila["Cliente"],
            fila["Val_Sol"],
            fila["IGV_Sol"],
            fila["Total_Sol"],
            fila["Val_Dol"],
            fila["IGV_Dol"],
            fila["Total_Dol"],
            "NO",
            "NO",
            "NA",
            "",
            fila["OC"],
            "XML — pendiente confirmar",
        ]
        for col_idx, valor in enumerate(valores, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.fill = fill_fila
            cell.border = borde
            cell.alignment = Alignment(vertical="center")
            if col_idx == 2 and valor:
                cell.number_format = "DD/MM/YYYY"

    anchos = [12, 14, 22, 16, 12, 12, 18, 12, 12, 10, 18, 20, 14, 15, 28]
    for i, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = ancho
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─── UI ────────────────────────────────────────────────────────────────────────

st.set_page_config(
    page_title="Control Nélida — NETSAT",
    layout="wide",
    page_icon="📊",
)

require_login()

st.markdown("<style>.block-container{padding-top:1.5rem}</style>", unsafe_allow_html=True)

col_title, col_reload = st.columns([5, 1])
with col_title:
    st.title("📊 Control de facturas y proyectos")
with col_reload:
    st.write("")
    if st.button("🔄 Recargar", use_container_width=True):
        cargar_proyectos.clear()
        cargar_guias.clear()
        cargar_facturas.clear()
        cargar_coupa.clear()
        parsear_xmls.clear()

fuente = "🌐 Drive Compartido" if os.path.exists(_PRIMARY) else "💾 local"
st.caption(f"Fuente: {fuente} · Actualizado por Nélida · {datetime.now().strftime('%d/%m/%Y %H:%M')}")

with st.spinner("Cargando..."):
    proyectos = cargar_proyectos()
    guias     = cargar_guias()
    facturas  = cargar_facturas()
    coupa     = cargar_coupa()
    xmls      = parsear_xmls()

# ─── MÉTRICAS ──────────────────────────────────────────────────────────────────
total_proy    = len(proyectos)
finalizados   = (proyectos.get("Estado", pd.Series()).astype(str).str.upper() == "FINALIZADO").sum()
en_curso      = total_proy - finalizados

fact_ant      = facturas[facturas["Cliente"].astype(str).str.upper() == "ANTAPACCAY"]
pendientes    = (fact_ant["Pagado"].astype(str).str.upper().isin(["NO", "NAN"])).sum()
total_usd     = fact_ant["Total_Dol"].sum()

coupa_pendiente = coupa[coupa["Fecha_Pago_Real"].isna()]["Monto"].sum()

c1, c2, c3, c4, c5, c6 = st.columns(6)
c1.metric("📁 Proyectos 2026",      total_proy)
c2.metric("✅ Finalizados",         finalizados)
c3.metric("🔄 En curso",            en_curso)
c4.metric("⏳ Fact. pendientes",    pendientes)
c5.metric("💵 Facturado Antap. USD", _usd(total_usd))
c6.metric("🏦 COUPA x cobrar USD",  _usd(coupa_pendiente))

st.divider()

# ─── TABS ──────────────────────────────────────────────────────────────────────
tab1, tab2, tab3, tab4, tab5 = st.tabs(["📁 Proyectos 2026", "📄 Guías", "🧾 Facturas", "🏦 COUPA", "💡 Sugerencias XML"])


# ══════════════════════════════════════════════════════════════
with tab1:
    f1, f2 = st.columns([2, 2])
    with f1:
        estados = ["Todos"] + sorted(proyectos["Estado"].dropna().unique().tolist()) if "Estado" in proyectos.columns else ["Todos"]
        filtro_estado = st.selectbox("Estado", estados)
    with f2:
        filtro_oc = st.text_input("Buscar por OC", placeholder="C003962782")

    vista = proyectos.copy()
    if filtro_estado != "Todos" and "Estado" in vista.columns:
        vista = vista[vista["Estado"] == filtro_estado]
    if filtro_oc and "OC_base" in vista.columns:
        vista = vista[vista["OC_base"].str.contains(filtro_oc, case=False, na=False)]

    cols = [c for c in ["Proyecto", "Cliente", "OC_base", "Factura_Venta", "Guia_Venta",
                         "Factura_Compra", "Estado", "Fecha_Inicio", "Fecha_Fin"] if c in vista.columns]
    vista_disp = vista[cols].copy()
    for col in ["Fecha_Inicio", "Fecha_Fin"]:
        if col in vista_disp.columns:
            vista_disp[col] = vista_disp[col].dt.strftime("%d/%m/%Y")

    st.caption(f"{len(vista_disp)} proyectos")
    st.dataframe(
        vista_disp,
        width="stretch",
        height=500,
        column_config={
            "Proyecto":       st.column_config.NumberColumn("N°",             width=60),
            "Cliente":        st.column_config.TextColumn("Cliente",          width=110),
            "OC_base":        st.column_config.TextColumn("OC",              width=130),
            "Factura_Venta":  st.column_config.TextColumn("Fact. Venta",     width=110),
            "Guia_Venta":     st.column_config.TextColumn("Guía",            width=110),
            "Factura_Compra": st.column_config.TextColumn("Fact. Compra",    width=110),
            "Estado":         st.column_config.TextColumn("Estado",          width=110),
            "Fecha_Inicio":   st.column_config.TextColumn("Inicio",          width=100),
            "Fecha_Fin":      st.column_config.TextColumn("Fin estimada",    width=100),
        },
        hide_index=True,
    )


# ══════════════════════════════════════════════════════════════
with tab2:
    f1, f2 = st.columns([2, 2])
    with f1:
        filtro_estatus = st.selectbox(
            "Estatus",
            ["Todos", "COMPLETO", "INCOMPLETO"],
        )
    with f2:
        filtro_guia = st.text_input("Buscar guía o OC", placeholder="T001-587 o C003962782")

    vg = guias.copy()
    if filtro_estatus != "Todos":
        vg = vg[vg["Estatus"].astype(str).str.upper() == filtro_estatus]
    if filtro_guia:
        mask = (
            vg["Guia"].astype(str).str.contains(filtro_guia, case=False, na=False) |
            vg["OC"].astype(str).str.contains(filtro_guia, case=False, na=False)
        )
        vg = vg[mask]

    vg_disp = vg[["Guia", "Fecha", "OC", "Estatus", "Factura", "Detalle"]].copy()
    vg_disp["Fecha"] = vg_disp["Fecha"].dt.strftime("%d/%m/%Y")

    completas   = (vg["Estatus"].astype(str).str.upper() == "COMPLETO").sum()
    incompletas = (vg["Estatus"].astype(str).str.upper() == "INCOMPLETO").sum()
    st.caption(f"{len(vg_disp)} guías · {completas} completas · {incompletas} incompletas")

    st.dataframe(
        vg_disp,
        width="stretch",
        height=500,
        column_config={
            "Guia":    st.column_config.TextColumn("Guía",    width=90),
            "Fecha":   st.column_config.TextColumn("Fecha",   width=100),
            "OC":      st.column_config.TextColumn("OC",      width=140),
            "Estatus": st.column_config.TextColumn("Estatus", width=100),
            "Factura": st.column_config.TextColumn("Factura", width=100),
            "Detalle": st.column_config.TextColumn("Detalle", width=400),
        },
        hide_index=True,
    )


# ══════════════════════════════════════════════════════════════
with tab3:
    f1, f2, f3 = st.columns([2, 2, 2])
    with f1:
        clientes = ["Todos"] + sorted(facturas["Cliente"].dropna().unique().tolist())
        filtro_cliente = st.selectbox("Cliente", clientes)
    with f2:
        filtro_pagado = st.selectbox("Pagado", ["Todos", "SI", "NO"])
    with f3:
        filtro_fact = st.text_input("Buscar factura u OC", placeholder="F001-480")

    vf = facturas.copy()
    if filtro_cliente != "Todos":
        vf = vf[vf["Cliente"] == filtro_cliente]
    if filtro_pagado != "Todos":
        vf = vf[vf["Pagado"].astype(str).str.upper() == filtro_pagado]
    if filtro_fact:
        mask = (
            vf["Factura"].astype(str).str.contains(filtro_fact, case=False, na=False) |
            vf["OC"].astype(str).str.contains(filtro_fact, case=False, na=False)
        )
        vf = vf[mask]

    vf_disp = vf[["Factura", "Fecha", "Cliente", "Total_Sol", "Total_Dol",
                   "Pagado", "Guia", "OC", "Observaciones"]].copy()
    vf_disp["Fecha"] = vf_disp["Fecha"].dt.strftime("%d/%m/%Y")

    total_sol_vis = vf["Total_Sol"].sum()
    total_dol_vis = vf["Total_Dol"].sum()
    st.caption(f"{len(vf_disp)} facturas · S/ {total_sol_vis:,.2f} · $ {total_dol_vis:,.2f}")

    st.dataframe(
        vf_disp,
        width="stretch",
        height=500,
        column_config={
            "Factura":      st.column_config.TextColumn("Factura",    width=110),
            "Fecha":        st.column_config.TextColumn("Fecha",      width=95),
            "Cliente":      st.column_config.TextColumn("Cliente",    width=160),
            "Total_Sol":    st.column_config.NumberColumn("Total S/", width=100, format="S/ %.2f"),
            "Total_Dol":    st.column_config.NumberColumn("Total $",  width=100, format="$ %.2f"),
            "Pagado":       st.column_config.TextColumn("Pagado",     width=75),
            "Guia":         st.column_config.TextColumn("Guía",       width=90),
            "OC":           st.column_config.TextColumn("OC",         width=130),
            "Observaciones":st.column_config.TextColumn("Obs.",       width=200),
        },
        hide_index=True,
    )


# ══════════════════════════════════════════════════════════════
with tab4:
    f1, f2 = st.columns([2, 2])
    with f1:
        filtro_coupa = st.selectbox(
            "Estado COUPA",
            ["Todos"] + sorted(coupa["COUPA"].dropna().unique().tolist()),
        )
    with f2:
        filtro_cobro = st.selectbox(
            "Cobrado",
            ["Todos", "Pendiente", "Cobrado"],
        )

    vc = coupa.copy()
    if filtro_coupa != "Todos":
        vc = vc[vc["COUPA"] == filtro_coupa]
    if filtro_cobro == "Pendiente":
        vc = vc[vc["Fecha_Pago_Real"].isna()]
    elif filtro_cobro == "Cobrado":
        vc = vc[vc["Fecha_Pago_Real"].notna()]

    vc_disp = vc[["Factura", "Fecha", "Monto", "Dias_Credito",
                   "Fecha_Pago_Pactada", "OC", "COUPA", "Fecha_Pago_Real", "Pago"]].copy()
    for col in ["Fecha", "Fecha_Pago_Pactada", "Fecha_Pago_Real"]:
        vc_disp[col] = vc_disp[col].dt.strftime("%d/%m/%Y")

    total_coupa    = vc["Monto"].sum()
    por_cobrar     = vc[vc["Fecha_Pago_Real"].isna()]["Monto"].sum()
    cobrado        = vc[vc["Fecha_Pago_Real"].notna()]["Monto"].sum()
    st.caption(f"{len(vc_disp)} facturas · $ {total_coupa:,.2f} total · $ {cobrado:,.2f} cobrado · $ {por_cobrar:,.2f} pendiente")

    st.dataframe(
        vc_disp,
        width="stretch",
        height=500,
        column_config={
            "Factura":            st.column_config.TextColumn("Factura",         width=110),
            "Fecha":              st.column_config.TextColumn("Emisión",         width=95),
            "Monto":              st.column_config.NumberColumn("Monto $",       width=100, format="$ %.2f"),
            "Dias_Credito":       st.column_config.TextColumn("Crédito",         width=90),
            "Fecha_Pago_Pactada": st.column_config.TextColumn("Vence",          width=95),
            "OC":                 st.column_config.TextColumn("OC",             width=130),
            "COUPA":              st.column_config.TextColumn("COUPA",          width=90),
            "Fecha_Pago_Real":    st.column_config.TextColumn("Fecha cobro",    width=100),
            "Pago":               st.column_config.TextColumn("Tipo",           width=70),
        },
        hide_index=True,
    )


# ══════════════════════════════════════════════════════════════
with tab5:
    if xmls.empty:
        st.warning("No se encontró la carpeta de XMLs en Z:. Verifica que la red esté conectada.")
    else:
        facturas_registradas = set(facturas["Factura"].astype(str).str.upper())
        nuevas = xmls[~xmls["Factura"].str.upper().isin(facturas_registradas)].copy()

        col_info1, col_info2, col_info3 = st.columns(3)
        col_info1.metric("📂 XMLs encontrados",        len(xmls))
        col_info2.metric("✅ Ya en archivo Nélida",    len(xmls) - len(nuevas))
        col_info3.metric("🆕 Nuevas (sin registrar)",  len(nuevas))

        if nuevas.empty:
            st.success("¡Todo al día! Todas las facturas XML ya están registradas en el archivo de Nélida.")
        else:
            st.info(
                f"Se encontraron **{len(nuevas)} facturas** en los XMLs que aún no aparecen "
                f"en el archivo de Nélida. Descarga el Excel de sugerencias, revísalo y agrégalo como referencia."
            )

            nuevas_disp = nuevas[["Factura", "Fecha", "Cliente", "Val_Dol", "IGV_Dol", "Total_Dol",
                                   "Val_Sol", "Total_Sol", "OC", "Moneda"]].copy()
            nuevas_disp["Fecha"] = nuevas_disp["Fecha"].dt.strftime("%d/%m/%Y")

            st.dataframe(
                nuevas_disp,
                width="stretch",
                height=400,
                column_config={
                    "Factura":   st.column_config.TextColumn("Factura",         width=110),
                    "Fecha":     st.column_config.TextColumn("Fecha",           width=100),
                    "Cliente":   st.column_config.TextColumn("Cliente",         width=150),
                    "Val_Dol":   st.column_config.NumberColumn("Valor $ s/IGV", width=120, format="$ %.2f"),
                    "IGV_Dol":   st.column_config.NumberColumn("IGV $",         width=95,  format="$ %.2f"),
                    "Total_Dol": st.column_config.NumberColumn("Total $",       width=100, format="$ %.2f"),
                    "Val_Sol":   st.column_config.NumberColumn("Valor S/",      width=100, format="S/ %.2f"),
                    "Total_Sol": st.column_config.NumberColumn("Total S/",      width=100, format="S/ %.2f"),
                    "OC":        st.column_config.TextColumn("OC",              width=130),
                    "Moneda":    st.column_config.TextColumn("Moneda",          width=75),
                },
                hide_index=True,
            )

            nombre_archivo = f"sugerencias_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            buf = _sugerencias_excel(nuevas)
            st.download_button(
                label="⬇️ Descargar sugerencias.xlsx",
                data=buf,
                file_name=nombre_archivo,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
            st.caption(
                "El archivo descargado tiene el mismo formato que el Excel de Nélida. "
                "Columnas PAGADO, GUIA REMISION y RETENCION quedan en blanco para que Nélida las complete."
            )
