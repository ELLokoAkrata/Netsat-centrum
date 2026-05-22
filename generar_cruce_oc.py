import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Datos de Supabase (resultado del query anterior)
supabase_rows = [
    ('C003909910',1,'000000000001675037','CABLE FIBRA OPTICA;LG 1 M,PANDUIT'),
    ('C003932633',1,'000000000001675939','RADIO; MUSICAL, PIONEER, MVH-S325BT'),
    ('C003948580',1,'000000000001679439','RADIO; MODELO: R7, CANALES: 1000'),
    ('C003973593',1,'000000000001675939','RADIO; MUSICAL, PIONEER, MVH-S325BT'),
    ('C003996089',1,'000000000001675176','ADAPTADOR COMUNIC;MUFA HORIZONTAL,IP68'),
    ('C003997749',1,'000000000001679439','RADIO; MODELO: R7, CANALES: 1000'),
    ('C004006358',1,None,'Base BOSH para Montaje de Camara PTZ'),
    ('C004018882',1,'000000000001675939','RADIO; MUSICAL, PIONEER, MVH-S325BT'),
    ('C004051881',1,'000000000001599931','ANTENA;KENWD,KRA-25'),
    ('C004058643',1,'000000000001778001','ANTENNA,RADIO;DSH,25 DBI,DOUBLE LINEAR'),
    ('C004059848',2,'000000000001778020','ADAPTER,COMM;FBROPTC,LC DUPLEX,LC'),
    ('C004065399',1,'000000000001543508','RADIO;136-174 MHZ, COMUNICACION PORTATIL'),
    ('C004078040',1,'000000000001674536','ADAPTADOR COMUNIC;LENV,4X91H17795,USB-C'),
    ('C004092257',1,'000000000001675052','MANGUITO ADAPT; AFL HYPERSCALE'),
    ('C004095571',1,'000000000001675037','CABLE FIBRA OPTICA;LG 1 M,PANDUIT'),
    ('C004099699',1,None,'Conector Recto 1" Liquid-tight, ABB 5334-TB'),
    ('C004099699',2,None,'Abrazaderas DIN 1", HYDAC HRL 5 A 33.7 PP UNC'),
    ('C004099699',3,None,'Fuente PoE MIDSPAN 60W VDC, AXIS T8144'),
    ('C004099699',4,None,'Cintillos 300mm PANDUIT PLT4H-L0'),
    ('C004099699',5,None,'Conversor voltaje 24Vdc MEAN WELL SD-50B-24'),
    ('C004099702',1,None,'Antena Omnidireccional 2.4GHz 6dBi MOBILEMARK OD6-2400MOD2-BLK'),
    ('C004099702',2,None,'Cintillos 300mm PANDUIT PLT4H-L0'),
    ('C004158379',1,'000000000001679439','RADIO; MODELO: R7'),
    ('C004158380',1,'000000000001543508','RADIO;136-174 MHZ, COMUNICACION PORTATIL'),
    ('C004158931',1,'000000000001331358','CARGADOR;MODULADOR IGBT ALPHA IGBT0125V035'),
    ('C004158931',2,'000000000001613805','CONTROLADOR; ALPHA TECHNOLOGIES 0180036-002'),
    ('C004158931',3,'000000000001685692','ANTENA RADIO;HELICOIDAL MOTOROLA PMAD4118'),
    ('C004158931',4,'000000000001685695','BATERIA RECAR;2200 MAH MOTOROLA PMNN4807'),
    ('C004158931',5,'000000000001674846','RECTIFICADOR; CORDEX 010-589-20-040'),
    ('C004158931',6,'000000000001674847','MODULO; VOLTAJE CORDEX 0180057'),
    ('C004158941',1,'000000000001256561','BATERIA RECAR;7.2V,1850 MAH,NIMH MOTOROLA HNN-9008'),
    ('C004174245',1,'000000000001273475','FUENTE ALIMNTCN;90/305 VAC MEAN WELL HLG-240H-15'),
    ('C004174248',1,'000000000001675939','RADIO; MUSICAL, PIONEER, MVH-S325BT'),
    ('C004174248',2,'000000000001679439','RADIO; MODELO: R7'),
    ('C004184303',1,None,'Jack Cat 6A PANDUIT CJ6X88TGBU'),
    ('C004184303',2,None,'Adaptador duplex SC PANDUIT CMDBUSCZBU'),
    ('C004184303',3,None,'Mufa FO tipo domo 8 puertos FIBERMAX CEV-D021'),
    ('C004184303',4,None,'Preformados 12.7mm 1/2 punto azul'),
    ('C004184306',1,None,'Bateria Gel 12Vdc 150Ah RITAR DG 12-150'),
    ('C004184306',2,None,'Conversor DC-DC MEAN WELL SD-150B-24'),
    ('C004188470',1,None,'Antena GPS Trimble Zephyr 3 Rugged 125000-30INT'),
    ('C004199390',1,'000000000001543508','RADIO;136-174 MHZ, COMUNICACION PORTATIL'),
    ('C004199396',1,'000000000001679439','RADIO; MODELO: R7'),
    ('C004204151',1,'000000000001679439','RADIO; TYPE: PORTABLE TRANSRECEIVER'),
    ('C004204153',1,'000000000001586550','CONECTOR CABLE;TIMES CONNECTOR TC-400-NMH-X'),
    ('C004204334',1,'000000000001795530','KIT;CUTTING BLADE Commscope CPT-BKS1'),
    ('C004204334',2,'000000000001795532','CONNECTOR,CBL;BNC Times EZ-400-BM-X'),
    ('C004204334',3,'000000000001795531','CABLE,COMM;COAXL Times LMR400'),
    ('C004204374',1,'000000000001795535','CONNECTOR,CBL;N,FEM Times TC-400-NFX'),
    ('C004204374',2,'000000000001795539','TIE,CBL;SS316 HELLERMANN 111-00295'),
    ('C004204374',3,'000000000001795538','CONNECTOR,CBL;BNC Amphenol 112603'),
    ('C004204374',4,'000000000001795533','CONNECTOR,CBL;BNC Times TC-400-BM-X'),
    ('C004204374',5,'000000000001795534','CONNECTOR,CBL;N,FEM Times EZ-400-NF-X'),
    ('C004204374',6,'000000000001795536','CONNECTOR,CBL;MINI UHF Times TC-400-MUHF'),
    ('C004204374',7,'000000000001795537','KIT;PREPARATION TOOLS Times TK-400EZ'),
    ('C004204427',1,'000000000001794685','TOOL;FLANGE HELLERMANN MK9SST'),
    ('C004204549',1,None,'Cuchilla de repuesto HELLERMANN 110-95273'),
    ('C004204550',1,'000000000001794687','TIE,CBL;BLK HELLERMANN 111-30001'),
    ('C004204557',1,'000000000001794688','TOOL;CABLE PREPARER Commscope CPT114U'),
    ('C004204559',1,'000000000001794689','TOOL;CABLE PREPARER Commscope CPT-L4ARC1'),
    ('C004211779',1,'000000000001257093','BATERIA RECAR;2100 MAH,NIMH MOTOROLA NTN9858C'),
    ('C004215039',1,'000000000001685698','MICROFONO;MOTOROLA PMMN4140'),
    ('C004215039',2,'000000000001732364','MICROFONO;PARLANTE REMOTO PMMN4013'),
    ('C004223441',2,None,'Cintas Brother TZE-661'),
    ('C004230888',1,'000000000001273528','BATERIA RECAR;IMPRES,3000 MAH MOTOROLA PMNN4493A'),
    ('C004230890',1,'000000000001543508','RADIO;136-174 MHZ, COMUNICACION PORTATIL'),
    ('C004230897',1,'000000000001675939','RADIO; MUSICAL, PIONEER, MVH-S325BT'),
    ('C004238626',1,'000000000001770293','SWITCH,NETWRK;PANDUIT CBXF12IW-AY 12 puertos'),
    ('C004238636',1,'000000000001770292','SWITCH,NETWRK;PANDUIT CBXF6IW-AY 6 puertos'),
]
supa_dict = {(r[0], r[1]): (r[2], r[3]) for r in supabase_rows}

# Leer Excel del padre
df_padre = pd.read_excel(
    'C:/Dev/Netsat-Centrum/OC Antapaccay pendientes de atencion (1).xlsx',
    sheet_name='Hoja1', header=1)
df_padre.columns = ['OC','Item','Desc_Papa','Cant','Fob','Desaduanaje','Impuestos','Venta_Unit','Venta_Total','Observaciones']
df_padre = df_padre[df_padre['OC'].astype(str).str.startswith('C00')].copy()
df_padre['Item'] = pd.to_numeric(df_padre['Item'], errors='coerce')
df_padre = df_padre.dropna(subset=['OC','Item']).drop_duplicates(subset=['OC','Item'], keep='first')
df_padre['Item'] = df_padre['Item'].astype(int)

# Construir filas del cruce
filas = []
for _, row in df_padre.iterrows():
    oc        = str(row['OC'])
    item      = int(row['Item'])
    desc_papa = str(row['Desc_Papa']) if pd.notna(row['Desc_Papa']) else ''
    codigo_oc, desc_oc = supa_dict.get((oc, item), (None, None))
    filas.append({
        'OC':                    oc,
        'ITEM':                  item,
        'DESCRIPCION (padre)':   desc_papa,
        'CODIGO (OC)':           codigo_oc or '',
        'DESCRIPCION (OC)':      desc_oc or '',
        'CODIGO (compra)':       '',
        'DESCRIPCION (compra)':  '',
    })

df_out = pd.DataFrame(filas)
ruta = 'C:/Dev/Netsat-Centrum/cruce_oc_compras.xlsx'
df_out.to_excel(ruta, index=False, sheet_name='Cruce')

# Formato visual
COLOR_PAPA   = 'BDD7EE'
COLOR_OC     = 'E2EFDA'
COLOR_COMPRA = 'FCE4D6'
GRIS_ENC     = 'D9D9D9'

wb = load_workbook(ruta)
ws = wb.active

# Insertar fila de grupos encima
ws.insert_rows(1)
ws.merge_cells('A1:C1'); ws['A1'] = 'EXCEL DE TU PAPA'
ws.merge_cells('D1:E1'); ws['D1'] = 'SALE DE LA OC'
ws.merge_cells('F1:G1'); ws['F1'] = 'OBJETIVO IDENTIFICAR EN LA COMPRA'

for col_idx, color, label in [(1, COLOR_PAPA, ''), (4, COLOR_OC, ''), (6, COLOR_COMPRA, '')]:
    c = ws.cell(1, col_idx)
    c.fill = PatternFill('solid', fgColor=color)
    c.font = Font(bold=True, size=11)
    c.alignment = Alignment(horizontal='center', vertical='center')

ws.row_dimensions[1].height = 22

# Encabezados de columna (fila 2)
seccion_color = {1: COLOR_PAPA, 2: COLOR_PAPA, 3: COLOR_PAPA,
                 4: COLOR_OC,   5: COLOR_OC,
                 6: COLOR_COMPRA, 7: COLOR_COMPRA}
for col in range(1, 8):
    c = ws.cell(2, col)
    c.fill = PatternFill('solid', fgColor=seccion_color[col])
    c.font = Font(bold=True)
    c.alignment = Alignment(horizontal='center', wrap_text=True)

# Filas de datos: alternar blanco/gris muy claro
for row_idx in range(3, ws.max_row + 1):
    bg = 'F2F2F2' if row_idx % 2 == 0 else 'FFFFFF'
    for col in range(1, 8):
        c = ws.cell(row_idx, col)
        c.fill = PatternFill('solid', fgColor=bg)
        c.alignment = Alignment(wrap_text=True, vertical='top')

# Anchos
for i, w in enumerate([15, 6, 42, 22, 42, 22, 42], 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = 'A3'

wb.save(ruta)
print('OK:', ruta)
print(f'Filas: {len(filas)} | Con codigo COUPA: {sum(1 for f in filas if f["CODIGO (OC)"])} | Sin codigo: {sum(1 for f in filas if not f["CODIGO (OC)"])}')
