import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font
from datetime import datetime
from auth import require_login

EXCEL_PATH = "OC_Antapaccay.xlsx"

@st.cache_data(ttl=0)
def cargar_datos():
    df = pd.read_excel(EXCEL_PATH, sheet_name="Hoja1", header=None)

    # Buscar la fila que contiene "OC" como encabezado
    header_row = 0
    for i, row in df.iterrows():
        if any(str(v).strip() == "OC" for v in row if pd.notna(v)):
            header_row = i
            break

    cols = df.iloc[header_row].tolist()
    df = df.iloc[header_row + 1:]          # sin reset_index: conserva posición original
    df.columns = [str(c).strip() if pd.notna(c) else f"_col{i}" for i, c in enumerate(cols)]

    # Mapeo flexible por nombre aproximado
    col_rename = {}
    targets = {
        "OC":            ["OC", "Nro OC", "N° OC", "Numero OC"],
        "Item":          ["Item", "Ítem", "N° Item", "Nro Item"],
        "Descripción":   ["Descripción", "Descripcion", "Descripción del producto", "Producto"],
        "Cant":          ["Cant", "Cantidad", "Qty"],
        "FOB":           ["FOB", "FOB US$", "Precio FOB"],
        "Desaduanaje":   ["Desaduanaje", "Desaduanamiento"],
        "Impuestos":     ["Impuestos", "Impuesto"],
        "Venta Unit US$":  ["Venta Unit US$", "Venta Unit US $", "PU Venta", "Precio Unit"],
        "Venta Total US$": ["Venta Total US$", "Venta Total US $", "Total Venta", "Total US$"],
        "Observaciones": ["Observaciones", "Obs", "Comentarios"],
    }
    for canonical, variants in targets.items():
        for v in variants:
            if v in df.columns:
                col_rename[v] = canonical
                break

    df = df.rename(columns=col_rename)
    for col in targets.keys():
        if col not in df.columns:
            df[col] = None
    df = df[list(targets.keys())]

    # Quitar filas completamente vacías — guardar fila Excel antes de filtrar
    df = df[df["OC"].notna() | df["Descripción"].notna()]
    df.insert(0, "Fila", df.index + 1)  # columna al inicio, Excel 1-based

    # Limpiar tipos
    for col in ["Cant", "FOB", "Desaduanaje", "Impuestos", "Venta Unit US$", "Venta Total US$"]:
        df[col] = pd.to_numeric(df[col], errors="coerce")
    df["OC"] = df["OC"].astype(str).str.strip().replace("nan", "")
    df["Item"] = df["Item"].astype(str).str.strip().replace("nan", "")
    df = df.reset_index(drop=True)
    return df

def guardar_excel(df):
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["Hoja1"]

    # Encontrar fila de encabezado
    header_row = None
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == "OC":
                header_row = cell.row
                break
        if header_row:
            break

    # Mapeo columnas Excel
    col_map = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value
        if val:
            col_map[str(val).strip()] = col

    fields = [
        ("OC", "OC"), ("Item", "Item"), ("Descripción", "Descripción"),
        ("Cant", "Cant"), ("FOB", "FOB"), ("Desaduanaje", "Desaduanaje"),
        ("Impuestos", "Impuestos"), ("Venta Unit US$", "Venta Unit US $"),
        ("Venta Total US$", "Venta Total US $"), ("Observaciones", "Observaciones"),
    ]

    next_row = ws.max_row + 1

    for _, row_data in df.iterrows():
        fila = row_data.get("Fila")
        excel_row = int(fila) if pd.notna(fila) else next_row
        if pd.isna(fila):
            next_row += 1
        for field, col_name in fields:
            col_idx = col_map.get(col_name) or col_map.get(field)
            if col_idx:
                val = row_data[field]
                ws.cell(row=excel_row, column=col_idx).value = None if pd.isna(val) else val

    wb.save(EXCEL_PATH)
    st.cache_data.clear()

# ─── UI ────────────────────────────────────────────────────────────────────────
st.set_page_config(page_title="OC Antapaccay — NETSAT", layout="wide", page_icon="📦")

require_login()

st.markdown("""
<style>
.block-container { padding-top: 1.5rem; }
.stDataFrame { font-size: 13px; }
[data-testid="stMetricValue"] { font-size: 1.6rem; font-weight: 600; }
</style>
""", unsafe_allow_html=True)

st.title("📦 OC Antapaccay — Pendientes de atención")
st.caption(f"Archivo: `{EXCEL_PATH}` · Última carga: {datetime.now().strftime('%d/%m/%Y %H:%M')}")

df = cargar_datos()

# ─── MÉTRICAS ──────────────────────────────────────────────────────────────────
col1, col2, col3 = st.columns(3)
total_items = len(df)
con_obs = df["Observaciones"].notna() & (df["Observaciones"].astype(str).str.strip() != "")
venta_total = df["Venta Total US$"].sum()

col1.metric("Total ítems", total_items)
col2.metric("📝 Con observaciones", int(con_obs.sum()))
col3.metric("💰 Venta total", f"${venta_total:,.0f}")

st.divider()

# ─── TABS ──────────────────────────────────────────────────────────────────────
tab1, tab2, tab3 = st.tabs(["📋 Ver & Editar", "➕ Agregar OC", "📊 Resumen por OC"])

# ══════════════════════════════════════════════════════════════
with tab1:
    fc1, fc2 = st.columns([2, 3])
    with fc1:
        filtro_oc = st.text_input("Buscar OC", placeholder="C004078040...")
    with fc2:
        filtro_desc = st.text_input("Buscar descripción / observaciones", placeholder="Motorola, Panduit...")

    df_view = df.copy()
    if filtro_oc:
        df_view = df_view[df_view["OC"].str.contains(filtro_oc, case=False, na=False)]
    if filtro_desc:
        mask = (
            df_view["Descripción"].str.contains(filtro_desc, case=False, na=False) |
            df_view["Observaciones"].str.contains(filtro_desc, case=False, na=False)
        )
        df_view = df_view[mask]

    st.caption(f"Mostrando {len(df_view)} de {total_items} ítems")

    edited = st.data_editor(
        df_view,
        width="stretch",
        height=520,
        column_config={
            "Fila":          st.column_config.NumberColumn("Fila", width=55, disabled=True),
            "OC":            st.column_config.TextColumn("OC", width=120),
            "Item":          st.column_config.TextColumn("Item", width=50),
            "Descripción":   st.column_config.TextColumn("Descripción", width=280),
            "Cant":          st.column_config.NumberColumn("Cant", width=60, format="%d"),
            "FOB":           st.column_config.NumberColumn("FOB US$", width=90, format="%.2f"),
            "Desaduanaje":   st.column_config.NumberColumn("Desaduanaje", width=100, format="%.2f"),
            "Impuestos":     st.column_config.NumberColumn("Impuestos", width=90, format="%.2f"),
            "Venta Unit US$":  st.column_config.NumberColumn("Unit US$", width=90, format="%.2f"),
            "Venta Total US$": st.column_config.NumberColumn("Total US$", width=100, format="%.2f"),
            "Observaciones": st.column_config.TextColumn("Observaciones", width=260),
        },
        num_rows="fixed",
        hide_index=True,
        key="editor_main",
    )

    if st.button("💾 Guardar cambios", type="primary"):
        df_updated = df.copy()
        for excel_row, row in edited.iterrows():
            if excel_row in df_updated.index:
                for col in edited.columns:
                    df_updated.at[excel_row, col] = row[col]
        guardar_excel(df_updated)
        st.success("✅ Guardado correctamente en el Excel")
        st.rerun()

# ══════════════════════════════════════════════════════════════
with tab2:
    st.subheader("Agregar nueva OC")
    with st.form("form_nuevo"):
        c1, c2 = st.columns([3, 1])
        oc_nuevo  = c1.text_input("Número de OC", placeholder="C004000000")
        item_nuevo = c2.text_input("Ítem", placeholder="1")

        desc_nuevo = st.text_input("Descripción completa del producto")

        c3, c4, c5 = st.columns(3)
        cant_nuevo = c3.number_input("Cantidad", min_value=0, step=1)
        fob_nuevo  = c4.number_input("FOB US$", min_value=0.0, format="%.2f")
        unit_nuevo = c5.number_input("Venta Unit US$", min_value=0.0, format="%.2f")

        obs_nuevo = st.text_input("Observaciones")

        submitted = st.form_submit_button("➕ Agregar OC", type="primary")
        if submitted:
            if not desc_nuevo:
                st.error("La descripción es obligatoria.")
            else:
                total_nuevo = cant_nuevo * unit_nuevo if cant_nuevo and unit_nuevo else None
                nueva_fila = {
                    "Fila": None,
                    "OC": oc_nuevo, "Item": item_nuevo, "Descripción": desc_nuevo,
                    "Cant": cant_nuevo or None, "FOB": fob_nuevo or None,
                    "Desaduanaje": None, "Impuestos": None,
                    "Venta Unit US$": unit_nuevo or None,
                    "Venta Total US$": total_nuevo,
                    "Observaciones": obs_nuevo,
                }
                df_nuevo = pd.concat([df, pd.DataFrame([nueva_fila])], ignore_index=True)
                guardar_excel(df_nuevo)
                st.success(f"✅ OC '{oc_nuevo}' agregada correctamente.")
                st.rerun()

# ══════════════════════════════════════════════════════════════
with tab3:
    st.subheader("Detalle por OC")
    por_oc = df[df["OC"] != ""].groupby("OC").agg(
        Ítems=("Item", "count"),
        Descripción=("Descripción", lambda x: x.iloc[0][:50] + "…" if len(x) > 0 else ""),
        Total_USD=("Venta Total US$", "sum"),
    ).reset_index().sort_values("Total_USD", ascending=False)

    st.dataframe(
        por_oc,
        width="stretch",
        column_config={
            "Total_USD": st.column_config.NumberColumn("Total US$", format="$%.2f"),
        },
        hide_index=True,
    )
